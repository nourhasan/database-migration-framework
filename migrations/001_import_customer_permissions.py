"""
Migration: Import Customer Permissions from Excel

This migration demonstrates:
- Reading an Excel file with multiple sheets
- Resolving database IDs from reference data
- Performing idempotent inserts
- Database-agnostic implementation
- Self-contained execution with transaction management

Excel Structure:
- Sheet 1 (Users): email
- Sheet 2 (Roles): role_name
- Sheet 3 (Customers): customer_name
- Rows are positionally aligned across sheets

Database Tables Expected:
- users (id, email)
- roles (id, name)
- customers (id, name)
- customer_permissions (user_id, role_id, customer_id)
"""

import sys
from pathlib import Path

# Add project root to path for imports
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from config.env_config import Config, setup_logging
from lib.db_connection import get_db_connection

# Set up logging for this migration
logger = setup_logging(__name__)


def read_excel_data(excel_path: str) -> list[dict]:
    """
    Read data from Excel file with three sheets: Users, Roles, Customers.
    
    Args:
        excel_path: Path to Excel file
    
    Returns:
        List of dictionaries with keys: email, role_name, customer_name
    
    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If required sheets are missing
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required to read Excel files. "
            "Install it with: pip install openpyxl"
        )
    
    logger.info(f"Reading Excel file: {excel_path}")
    
    if not Path(excel_path).exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    
    # Verify required sheets exist
    required_sheets = ['Users', 'Roles', 'Customers']
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Required sheet '{sheet_name}' not found in Excel file. "
                f"Available sheets: {workbook.sheetnames}"
            )
    
    # Read data from each sheet (skip header row)
    users_sheet = workbook['Users']
    roles_sheet = workbook['Roles']
    customers_sheet = workbook['Customers']
    
    # Extract data starting from row 2 (row 1 is header)
    users = [row[0].value for row in users_sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
    roles = [row[0].value for row in roles_sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
    customers = [row[0].value for row in customers_sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
    
    # Ensure all lists have the same length
    min_length = min(len(users), len(roles), len(customers))
    if len(users) != len(roles) or len(roles) != len(customers):
        logger.warning(
            f"Sheet row counts differ: Users={len(users)}, Roles={len(roles)}, "
            f"Customers={len(customers)}. Using minimum length: {min_length}"
        )
    
    # Combine data from all three sheets
    data = []
    for i in range(min_length):
        data.append({
            'email': users[i],
            'role_name': roles[i],
            'customer_name': customers[i]
        })
    
    logger.info(f"Read {len(data)} rows from Excel file")
    
    return data


def resolve_user_id(db, email: str) -> int:
    """
    Resolve user ID from email address.
    
    Args:
        db: Database connection
        email: User email address
    
    Returns:
        User ID
    
    Raises:
        ValueError: If user not found
    """
    db.execute("SELECT id FROM users WHERE email = ?", (email,))
    result = db.fetchone()
    
    if not result:
        raise ValueError(f"User not found: {email}")
    
    return result[0]


def resolve_role_id(db, role_name: str) -> int:
    """
    Resolve role ID from role name.
    
    Args:
        db: Database connection
        role_name: Role name
    
    Returns:
        Role ID
    
    Raises:
        ValueError: If role not found
    """
    db.execute("SELECT id FROM roles WHERE name = ?", (role_name,))
    result = db.fetchone()
    
    if not result:
        raise ValueError(f"Role not found: {role_name}")
    
    return result[0]


def resolve_customer_id(db, customer_name: str) -> int:
    """
    Resolve customer ID from customer name.
    
    Args:
        db: Database connection
        customer_name: Customer name
    
    Returns:
        Customer ID
    
    Raises:
        ValueError: If customer not found
    """
    db.execute("SELECT id FROM customers WHERE name = ?", (customer_name,))
    result = db.fetchone()
    
    if not result:
        raise ValueError(f"Customer not found: {customer_name}")
    
    return result[0]


def permission_exists(db, user_id: int, role_id: int, customer_id: int) -> bool:
    """
    Check if a customer permission already exists.
    
    Args:
        db: Database connection
        user_id: User ID
        role_id: Role ID
        customer_id: Customer ID
    
    Returns:
        True if permission exists, False otherwise
    """
    db.execute(
        """
        SELECT COUNT(*) 
        FROM customer_permissions 
        WHERE user_id = ? AND role_id = ? AND customer_id = ?
        """,
        (user_id, role_id, customer_id)
    )
    
    count = db.fetchone()[0]
    return count > 0


def insert_permission(db, user_id: int, role_id: int, customer_id: int):
    """
    Insert a customer permission record (idempotent).
    
    Args:
        db: Database connection
        user_id: User ID
        role_id: Role ID
        customer_id: Customer ID
    """
    # Check if permission already exists
    if permission_exists(db, user_id, role_id, customer_id):
        logger.debug(
            f"Permission already exists: user_id={user_id}, "
            f"role_id={role_id}, customer_id={customer_id}"
        )
        return
    
    # Insert new permission
    db.execute(
        """
        INSERT INTO customer_permissions (user_id, role_id, customer_id, created_at)
        VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        """,
        (user_id, role_id, customer_id)
    )
    
    logger.debug(
        f"Inserted permission: user_id={user_id}, "
        f"role_id={role_id}, customer_id={customer_id}"
    )


def run_migration():
    """
    Execute the customer permissions import migration.
    
    This function:
    1. Reads Excel data
    2. Opens a database connection
    3. Resolves IDs for users, roles, and customers
    4. Inserts permissions (idempotent)
    5. Commits on success or rolls back on failure
    """
    logger.info("=" * 70)
    logger.info("Starting migration: Import Customer Permissions")
    logger.info("=" * 70)
    
    db = None
    
    try:
        # Validate configuration
        Config.validate()
        
        # Get Excel file path from configuration
        excel_path = Config.EXCEL_PATH
        if not excel_path:
            raise ValueError(
                "EXCEL_PATH not set in environment configuration. "
                "Please set it in your .env file."
            )
        
        # Read Excel data
        data = read_excel_data(excel_path)
        
        if not data:
            logger.warning("No data found in Excel file. Nothing to import.")
            return
        
        # Open database connection
        logger.info(f"Connecting to database: {Config.DB_ENGINE}")
        db = get_db_connection()
        
        # Process each row
        inserted_count = 0
        skipped_count = 0
        
        for idx, row in enumerate(data, start=1):
            try:
                # Resolve IDs
                user_id = resolve_user_id(db, row['email'])
                role_id = resolve_role_id(db, row['role_name'])
                customer_id = resolve_customer_id(db, row['customer_name'])
                
                # Check if permission exists before insert
                if permission_exists(db, user_id, role_id, customer_id):
                    skipped_count += 1
                    logger.debug(f"Row {idx}: Permission already exists, skipping")
                else:
                    # Insert permission
                    insert_permission(db, user_id, role_id, customer_id)
                    inserted_count += 1
                    logger.info(
                        f"Row {idx}: Inserted permission for {row['email']} "
                        f"({row['role_name']}) on {row['customer_name']}"
                    )
                
            except ValueError as e:
                # Log the error but continue processing other rows
                logger.error(f"Row {idx}: Skipping due to error: {e}")
                continue
        
        # Commit transaction
        db.commit()
        
        logger.info("=" * 70)
        logger.info(f"Migration completed successfully")
        logger.info(f"Total rows processed: {len(data)}")
        logger.info(f"New permissions inserted: {inserted_count}")
        logger.info(f"Existing permissions skipped: {skipped_count}")
        logger.info("=" * 70)
        
    except Exception as e:
        logger.error(f"Migration failed: {e}")
        
        # Rollback transaction on error
        if db:
            db.rollback()
            logger.info("Transaction rolled back")
        
        # Re-raise the exception
        raise
    
    finally:
        # Always close the database connection
        if db:
            db.close()
            logger.info("Database connection closed")


if __name__ == '__main__':
    try:
        run_migration()
    except Exception as e:
        logger.error(f"Migration execution failed: {e}")
        sys.exit(1)
